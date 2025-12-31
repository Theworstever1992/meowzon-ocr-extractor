"""
Output handling for multiple file formats
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional

from .logging_utils import LoggerMixin


class OutputHandler(LoggerMixin):
    """Handle multiple output formats"""
    
    @staticmethod
    def save_csv(df: pd.DataFrame, path: str):
        """Save dataframe as CSV"""
        df.to_csv(path, index=False, encoding='utf-8-sig')
    
    @staticmethod
    def save_excel(df: pd.DataFrame, path: str):
        """Save dataframe as formatted Excel file"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        # Write to Excel
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Orders')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Orders']
            
            # Style header row
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                column_letter = worksheet.cell(row=1, column=idx).column_letter
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            worksheet.freeze_panes = 'A2'
    
    @staticmethod
    def save_json(df: pd.DataFrame, path: str):
        """Save dataframe as JSON"""
        df.to_json(path, orient='records', indent=2, force_ascii=False)
    
    @staticmethod
    def save_html_report(
        df: pd.DataFrame,
        path: str,
        summary: Optional[dict] = None,
        analytics_plot: Optional[str] = None
    ):
        """
        Generate HTML report with statistics and data table
        
        Args:
            df: DataFrame with order data
            path: Output path for HTML file
            summary: Summary statistics dictionary
            analytics_plot: Path to analytics plot image
        """
        html_parts = []
        
        # HTML header
        html_parts.append("""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Meowzon Extraction Report</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 1400px;
            margin: 0 auto;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }
        .container {
            background: white;
            border-radius: 10px;
            padding: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        }
        h1 {
            color: #764ba2;
            text-align: center;
            font-size: 2.5em;
            margin-bottom: 10px;
        }
        .cat-emoji {
            font-size: 3em;
            text-align: center;
            margin-bottom: 20px;
        }
        .summary {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }
        .summary-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .summary-card h3 {
            margin: 0 0 10px 0;
            font-size: 0.9em;
            opacity: 0.9;
        }
        .summary-card .value {
            font-size: 2em;
            font-weight: bold;
        }
        .plot-container {
            text-align: center;
            margin: 30px 0;
        }
        .plot-container img {
            max-width: 100%;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 30px;
            font-size: 0.9em;
        }
        th {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 12px;
            text-align: left;
            position: sticky;
            top: 0;
        }
        td {
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }
        tr:hover {
            background: #f5f5f5;
        }
        .timestamp {
            text-align: center;
            color: #666;
            margin-top: 30px;
            font-size: 0.9em;
        }
        .status-success {
            color: #28a745;
            font-weight: bold;
        }
        .status-review {
            color: #ffc107;
            font-weight: bold;
        }
        .status-failed {
            color: #dc3545;
            font-weight: bold;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="cat-emoji">😺</div>
        <h1>Meowzon Extraction Report</h1>
""")
        
        # Summary statistics
        if summary:
            html_parts.append('<div class="summary">')
            
            cards = [
                ('Total Files', summary['total_files']),
                ('Successful', summary['successful_extractions']),
                ('Review Required', summary['review_required']),
                ('Total Orders', summary['total_orders']),
                ('Total Items', summary['total_items']),
                ('Avg Confidence', f"{summary['avg_confidence']:.1f}%"),
                ('AI Usage', f"{summary['ai_usage_rate']:.1f}%"),
                ('Unique Sellers', summary['unique_sellers']),
            ]
            
            for title, value in cards:
                html_parts.append(f'''
                <div class="summary-card">
                    <h3>{title}</h3>
                    <div class="value">{value}</div>
                </div>
                ''')
            
            html_parts.append('</div>')
        
        # Analytics plot
        if analytics_plot and Path(analytics_plot).exists():
            html_parts.append(f'''
            <div class="plot-container">
                <h2>Analytics</h2>
                <img src="{Path(analytics_plot).name}" alt="Analytics Plot">
            </div>
            ''')
        
        # Data table
        html_parts.append('<h2>Extraction Results</h2>')
        
        # Convert DataFrame to HTML with custom styling
        df_html = df.to_html(
            index=False,
            classes='data-table',
            escape=False,
            na_rep='-'
        )
        
        # Add status styling
        df_html = df_html.replace(
            'Success', '<span class="status-success">Success</span>'
        )
        df_html = df_html.replace(
            'Review', '<span class="status-review">Review</span>'
        )
        df_html = df_html.replace(
            'Failed', '<span class="status-failed">Failed</span>'
        )
        
        html_parts.append(df_html)
        
        # Timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        html_parts.append(f'''
        <div class="timestamp">
            Report generated on {timestamp}
        </div>
        ''')
        
        # HTML footer
        html_parts.append('''
    </div>
</body>
</html>
''')
        
        # Write to file
        with open(path, 'w', encoding='utf-8') as f:
            f.write(''.join(html_parts))
    
    @staticmethod
    def save_all_formats(
        df: pd.DataFrame,
        base_path: str,
        summary: Optional[dict] = None,
        analytics_plot: Optional[str] = None
    ):
        """
        Save dataframe in all supported formats
        
        Args:
            df: DataFrame to save
            base_path: Base path without extension
            summary: Summary statistics
            analytics_plot: Path to analytics plot
        """
        base = Path(base_path).stem
        directory = Path(base_path).parent
        
        formats = {
            'csv': OutputHandler.save_csv,
            'xlsx': OutputHandler.save_excel,
            'json': OutputHandler.save_json,
        }
        
        saved_files = []
        
        for ext, save_func in formats.items():
            try:
                output_path = directory / f"{base}.{ext}"
                save_func(df, str(output_path))
                saved_files.append(str(output_path))
            except Exception as e:
                print(f"Warning: Failed to save {ext}: {e}")
        
        # HTML report
        try:
            html_path = directory / f"{base}_report.html"
            OutputHandler.save_html_report(
                df, str(html_path), summary, analytics_plot
            )
            saved_files.append(str(html_path))
        except Exception as e:
            print(f"Warning: Failed to save HTML report: {e}")
        
        return saved_files


def format_output_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format dataframe for better readability in output
    
    Args:
        df: Raw dataframe
    
    Returns:
        Formatted dataframe
    """
    # Round confidence to 1 decimal
    if 'Tesseract Confidence (%)' in df.columns:
        df['Tesseract Confidence (%)'] = df['Tesseract Confidence (%)'].round(1)
    
    # Truncate long raw text
    if 'Raw Tesseract Snippet' in df.columns:
        df['Raw Tesseract Snippet'] = df['Raw Tesseract Snippet'].str[:200]
    
    # Sort by filename
    if 'File Name' in df.columns:
        df = df.sort_values('File Name')
    
    return df
